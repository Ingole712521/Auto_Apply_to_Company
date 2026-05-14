"""
Read contacts from Excel, fill {{COMPANY_NAME}} in a text email template, send via Gmail SMTP.

Excel: company name + recipient email. Recognized headers (case-insensitive, extra spaces OK):
  Email: Email, E-mail, EmailID, Email ID, Mail, Recipient, To
  Company: Company, Company Name, Comapany Name (common typo), Organization, Business

If row 1 is a title and row 2 has headers (like your sheet), use auto-detect (default) or --header-row 1.

Usage:
  copy .env.example to .env and set GMAIL_ADDRESS + GMAIL_APP_PASSWORD
  pip install -r requirements.txt
  python send_mail_merge.py --excel contacts.xlsx --template email_template.txt

Dry run (no sending):
  python send_mail_merge.py --excel contacts.xlsx --template email_template.txt --dry-run

By default attaches Nehal_Ingole_7397966719.pdf from the same folder as this script. Override with
  --attach path/to/resume.pdf   or send without file using   --no-attach

By default only the first 50 sheet rows (in file order) are processed. Use --no-limit for the full list
  or e.g. --limit 100 to change the cap.

After a row's email(s) send successfully, writes Done to the status column (column D by default). Use
  --no-mark-done to skip updating the workbook. --done-column E to pick another column letter.
"""

from __future__ import annotations

import argparse
import os
import re
import smtplib
import ssl
import time
from email.message import EmailMessage
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

load_dotenv()

EMAIL_ALIASES = {
    "email",
    "e-mail",
    "emailid",
    "email id",
    "email_id",
    "mail",
    "recipient",
    "to",
    "email address",
}
COMPANY_ALIASES = {
    "company",
    "company name",
    "comapany name",
    "organization",
    "business",
    "firm",
    "client",
}
STATUS_ALIASES = {
    "status",
    "sent",
    "state",
    "remark",
    "remarks",
    "done",
    "email status",
}

# Basic check: skip "NO email id", N/A, non-address text
_SIMPLE_EMAIL = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")


def normalize_header(s: str) -> str:
    return str(s).strip().lower()


def pick_column(df: pd.DataFrame, aliases: set[str]) -> str | None:
    mapping = {normalize_header(c): c for c in df.columns}
    for alias in aliases:
        if alias in mapping:
            return mapping[alias]
    return None


def is_valid_recipient_email(raw: str) -> bool:
    s = str(raw).strip()
    if not s or s.lower() == "nan":
        return False
    low = s.lower()
    if "no email" in low or low in {"n/a", "na", "-", "none", "tbd"}:
        return False
    if "@" not in s:
        return False
    return bool(_SIMPLE_EMAIL.match(s))


def parse_recipient_emails(raw: str) -> list[str]:
    """Split comma/semicolon-separated addresses; return unique valid emails in order."""
    s = str(raw).strip()
    if not s or s.lower() == "nan":
        return []
    low = s.lower()
    if "no email" in low or low in {"n/a", "na", "-", "none", "tbd"}:
        return []
    parts = re.split(r"[,;\n]+", s)
    out: list[str] = []
    seen: set[str] = set()
    for p in parts:
        e = p.strip()
        if not is_valid_recipient_email(e):
            continue
        k = e.lower()
        if k not in seen:
            seen.add(k)
            out.append(e)
    return out


def resolve_done_column_letter(df: pd.DataFrame, override: str | None) -> str:
    """Excel column letter where we write Done (default D, or a named status column if found)."""
    if override:
        return override.strip().upper()
    col = pick_column(df, STATUS_ALIASES)
    if col is not None:
        loc = df.columns.get_loc(col)
        if isinstance(loc, int):
            return get_column_letter(loc + 1)
    return "D"


def write_done_to_sheet(excel_path: Path, row_1based: int, col_letter: str, value: str = "Done") -> None:
    col_idx = column_index_from_string(col_letter)
    wb = load_workbook(excel_path)
    ws = wb.active
    ws.cell(row=row_1based, column=col_idx, value=value)
    wb.save(excel_path)


def load_contacts(path: Path, header_row: int | None) -> tuple[pd.DataFrame, int]:
    """Return (dataframe, header_row_index). header_row is 0-based pandas ``header=`` argument."""
    if header_row is not None:
        df = pd.read_excel(path, header=header_row)
        if pick_column(df, EMAIL_ALIASES) and pick_column(df, COMPANY_ALIASES):
            return df, header_row
        raise SystemExit(
            f"--header-row {header_row} did not yield recognizable columns. Found: {list(df.columns)}"
        )

    for header in range(0, 8):
        df = pd.read_excel(path, header=header)
        if pick_column(df, EMAIL_ALIASES) and pick_column(df, COMPANY_ALIASES):
            return df, header

    sample = pd.read_excel(path, header=0)
    raise SystemExit(
        "Could not find company + email columns. First row headers look like: "
        f"{list(sample.columns)}. Set --header-row to the 0-based row that contains "
        "'Comapany Name' / 'EmailID' (often 1 if row 1 is a title)."
    )


def parse_template(path: Path) -> tuple[str, str]:
    text = path.read_text(encoding="utf-8")
    subject = "Message"
    body = text
    m = re.match(r"^\s*SUBJECT:\s*(.+?)\s*\n", text, re.IGNORECASE | re.DOTALL)
    if m:
        subject = m.group(1).strip()
        body = text[m.end() :].lstrip("\n")
    return subject, body


def render(template: str, company: str) -> str:
    return template.replace("{{COMPANY_NAME}}", company)


def send_gmail(
    sender: str,
    app_password: str,
    to_addr: str,
    subject: str,
    body: str,
    attachment: Path | None,
) -> None:
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = to_addr
    msg.set_content(body)
    if attachment is not None:
        pdf_bytes = attachment.read_bytes()
        msg.add_attachment(
            pdf_bytes,
            maintype="application",
            subtype="pdf",
            filename=attachment.name,
        )

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(sender, app_password.replace(" ", ""))
        server.send_message(msg)


def main() -> None:
    parser = argparse.ArgumentParser(description="Mail merge from Excel via Gmail")
    parser.add_argument("--excel", type=Path, default=Path("contacts.xlsx"), help="Path to .xlsx")
    parser.add_argument(
        "--template",
        type=Path,
        default=Path("email_template.txt"),
        help="Plain text template; first line may be SUBJECT: ...; body uses {{COMPANY_NAME}}",
    )
    parser.add_argument("--dry-run", action="store_true", help="Print actions without sending")
    parser.add_argument("--delay", type=float, default=2.0, help="Seconds between sends (rate limit)")
    parser.add_argument(
        "--limit",
        type=int,
        default=50,
        metavar="N",
        help="Process at most the first N rows from the sheet in order (default: 50). Ignored with --no-limit.",
    )
    parser.add_argument(
        "--no-limit",
        action="store_true",
        help="Process all rows in the sheet (ignore --limit).",
    )
    parser.add_argument(
        "--header-row",
        type=int,
        default=None,
        metavar="N",
        help="0-based Excel row index used as column names (default: try rows 0–7; use 1 if headers are on sheet row 2)",
    )
    parser.add_argument(
        "--attach",
        type=Path,
        default=None,
        metavar="FILE.pdf",
        help="PDF resume to attach (default: Nehal_Ingole_7397966719.pdf next to this script)",
    )
    parser.add_argument(
        "--no-attach",
        action="store_true",
        help="Send email without any attachment",
    )
    parser.add_argument(
        "--done-column",
        default=None,
        metavar="LETTER",
        help="Excel column letter to write Done after successful sends (default: D, or first matching Status/Sent column)",
    )
    parser.add_argument(
        "--no-mark-done",
        action="store_true",
        help="Do not update the Excel file after sends",
    )
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    if args.no_attach:
        attach_path: Path | None = None
    elif args.attach is not None:
        attach_path = args.attach.expanduser()
        if not attach_path.is_file():
            raise SystemExit(f"Attachment not found: {attach_path.resolve()}")
        attach_path = attach_path.resolve()
    else:
        default_pdf = script_dir / "Nehal_Ingole_7397966719.pdf"
        attach_path = default_pdf.resolve() if default_pdf.is_file() else None

    sender = os.environ.get("GMAIL_ADDRESS", "").strip()
    app_password = os.environ.get("GMAIL_APP_PASSWORD", "").strip()
    env_subject = os.environ.get("EMAIL_SUBJECT", "").strip()

    if not args.excel.is_file():
        raise SystemExit(f"Excel file not found: {args.excel.resolve()}")

    if not args.template.is_file():
        raise SystemExit(f"Template not found: {args.template.resolve()}")

    subject_tpl, body_tpl = parse_template(args.template)

    df, header_used = load_contacts(args.excel, args.header_row)
    if df.empty:
        raise SystemExit("Excel file has no data rows under the detected header.")

    email_col = pick_column(df, EMAIL_ALIASES)
    company_col = pick_column(df, COMPANY_ALIASES)
    if not email_col:
        raise SystemExit(
            f"Could not find an email column. Headers: {list(df.columns)}. "
            f"Use one of: {sorted(EMAIL_ALIASES)}"
        )
    if not company_col:
        raise SystemExit(
            f"Could not find a company column. Headers: {list(df.columns)}. "
            f"Use one of: {sorted(COMPANY_ALIASES)}"
        )

    total_rows = len(df)
    if not args.no_limit:
        if args.limit < 1:
            raise SystemExit("--limit must be at least 1, or use --no-limit for the full sheet.")
        df = df.head(args.limit)
        print(
            f"Row cap: first {len(df)} of {total_rows} data row(s) (--limit {args.limit}). "
            "Use --no-limit to process everyone."
        )
    else:
        print(f"Row cap: disabled (--no-limit); processing all {len(df)} data row(s).")

    done_col_letter = resolve_done_column_letter(df, args.done_column)
    mark_done = not args.no_mark_done
    if mark_done:
        print(f"Will write 'Done' to column {done_col_letter} after each fully successful row (use --no-mark-done to disable).")

    if not args.dry_run:
        if not sender or not app_password:
            raise SystemExit("Set GMAIL_ADDRESS and GMAIL_APP_PASSWORD in .env (see .env.example).")
        if attach_path is None and not args.no_attach:
            raise SystemExit(
                f"No PDF attachment: place Nehal_Ingole_7397966719.pdf in {script_dir} "
                "or pass --attach path/to/resume.pdf (or use --no-attach to send without a file)."
            )

    sent = 0
    for pos, (idx, row) in enumerate(df.iterrows()):
        company = str(row[company_col]).strip()
        to_raw = str(row[email_col]).strip()
        excel_row_1based = header_used + 2 + pos
        if not company or company.lower() == "nan":
            print(f"Sheet ~row {excel_row_1based}: skip — empty company name")
            continue
        addresses = parse_recipient_emails(to_raw)
        if not addresses:
            print(
                f"Sheet ~row {excel_row_1based}: skip — no valid email in ({to_raw!r}) for {company!r}"
            )
            continue

        subj = env_subject or subject_tpl
        subj = render(subj, company)
        body = render(body_tpl, company)

        if args.dry_run:
            print(f"[dry-run] Recipients: {', '.join(addresses)} | Company: {company}")
            print(f"  Subject: {subj}")
            if attach_path is not None:
                print(f"  Attachment: {attach_path.name} ({attach_path})")
            else:
                print("  Attachment: (none)")
            if mark_done:
                print(f"  Would mark row {excel_row_1based} column {done_col_letter} = Done")
            print("  --- body preview ---")
            print(body[:500] + ("..." if len(body) > 500 else ""))
            print()
        else:
            row_ok = True
            for addr in addresses:
                try:
                    send_gmail(sender, app_password, addr, subj, body, attach_path)
                    sent += 1
                    print(f"Sent to {addr} ({company})")
                    if args.delay > 0:
                        time.sleep(args.delay)
                except (smtplib.SMTPException, OSError) as exc:
                    row_ok = False
                    print(f"Sheet row {excel_row_1based}: failed sending to {addr!r}: {exc}")
                    break
            if row_ok and mark_done:
                try:
                    write_done_to_sheet(args.excel.resolve(), excel_row_1based, done_col_letter)
                    print(f"  Marked sheet row {excel_row_1based} column {done_col_letter} = Done")
                except PermissionError:
                    print(
                        f"  Could not save Excel (is {args.excel.name} open in Excel?). "
                        "Close it and re-run this row or mark Done manually."
                    )
                except Exception as exc:
                    print(f"  Could not update Excel row {excel_row_1based}: {exc}")

    if args.dry_run:
        print("Dry run complete; no messages were sent.")
    else:
        print(f"Done. Sent {sent} message(s).")


if __name__ == "__main__":
    main()
